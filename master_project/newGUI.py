#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Nov 19 14:19:53 2021

@author: Amir Ershadi
"""

import tkinter as Tk
import os
import numpy as np
import pandas as pd
import time
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk)
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure
from numpy import loadtxt
from tkinter import *
from tkinter.filedialog import askopenfile 
import tkinter as tk
import platform
import matplotlib.pyplot as plt
import openpyxl    
import xlsxwriter     
global mainw
from matplotlib import cycler
global filename
import math



# Default plot style
colors = cycler('color',
    ['#EE6666', '#3388BB', '#9988DD',
     '#EECC55', '#88BB44', '#FFBBBB'])
plt.rc('axes', facecolor='#E6E6E6', edgecolor='none',
axisbelow=True, grid=True, prop_cycle=colors)
plt.rc('xtick', direction='out', color='black')
plt.rc('ytick', direction='out', color='black')
plt.rc('patch', edgecolor='#E6E6E6')
plt.rc('lines', linewidth=2)
#################################################
#Tkinter window
win=Tk()
screen_width=win.winfo_screenwidth()
screen_height=win.winfo_screenheight()
title = win.title("Calibration_GUI")
win.geometry(f'{screen_width}x{screen_height}')
frame_canvas = tk.Frame(win)
frame_canvas.grid(row=0, column=0, pady=(0, 0), sticky='w')
frame_canvas.grid_rowconfigure(0, weight=1)
frame_canvas.grid_columnconfigure(0, weight=1)
# Set grid_propagate to False to allow 5-by-5 buttons resizing later
frame_canvas.grid_propagate(False)
canvasf = tk.Canvas(frame_canvas)
canvasf.grid(row=0, column=0, sticky="news")
vsb = tk.Scrollbar(frame_canvas, orient="vertical", command=canvasf.yview)
vsb.grid(row=0, column=2, sticky='ns')
vsbb = tk.Scrollbar(frame_canvas, orient="horizontal", command=canvasf.xview)
vsbb.grid(row=0, column=1)
canvasf.configure(yscrollcommand=vsb.set)
canvasf.configure(xscrollcommand=vsbb.set)
mainw = tk.Frame(canvasf)
canvasf.create_window((0, 0), window=mainw, anchor='nw')
# Update buttons frames idle tasks to let tkinter calculate buttons sizes
mainw.update_idletasks()
frame_canvas.config(width=screen_width-10,
                    height=screen_height-10)

# Set the canvas scrolling region
canvasf.config(scrollregion=canvasf.bbox("all"))
global mag
global calib_txt
global sheet
global extra
#Load data
calib_txt= open("calibration.txt", encoding="utf8", errors='ignore')
calib_txt = calib_txt.read()
#Find the path
path=os.path.abspath('calibration.txt')
if platform.system()=='Darwin':
    path=path.replace("/calibration.txt", "")
if platform.system()=='Linux':
    path=path.replace("/calibration.txt", "")    
if platform.system()=='Windows':
    path=path.replace("\calibration.txt", "")    
wb = openpyxl.load_workbook('calibration.xlsx')
sheet = wb["Sheet1"]
calib_txt=calib_txt.split ("\t")
disc = LabelFrame(mainw, text=calib_txt[0], font="Arial 12 bold italic")
disc.grid(row=0,column=0)
scrollbar = Scrollbar(disc)
discrip = tk.Text(disc, height=7, width=70, yscrollcommand=scrollbar.set, font="Arial 14")
scrollbar.config(command=discrip.yview)
scrollbar.pack(side=RIGHT, fill=Y)
discrip.pack(side="left")
discrip.insert(END,calib_txt[1])  
opt = LabelFrame(mainw , text="Calibration options:", font="Arial 10 bold italic")
opt.grid(row=1,column=0, sticky='W')   
calibration_option = StringVar(mainw)
calibration_option.set(' Calibration type ')
option_choose = OptionMenu(opt, calibration_option, *sorted({'Magnetometer sensor'\
        , 'Gyro', 'Arm radius step1','Arm radius step2', 'Inclination angle'}),)
option_choose.grid(row=0, column=0, sticky='W', padx=5, pady=2)
check_butt = Button(master = opt, command =lambda:run_filter(), height = 1, width =2, text = "☑️") 
check_butt.grid(row=0,column=1)
manual_butt = Button(master = opt, command =lambda:manual(), height = 1, width =10, text = "READ MANUAL",fg='red') 
manual_butt.grid(row=0,column=2)
#Open PDF manual
if platform.system()=='Darwin':
    def manual():
        import subprocess
        subprocess.call(['open','-a','Preview','readme.pdf'])
def clean():
        try:
            for widget in extra.winfo_children():
                    widget.destroy() 
        except:
            pass
        try:
            for widget in extra1.winfo_children():
                    widget.destroy() 
        except:
            pass  
##########################MAGNETOMETER CALIBRATION############################        
def run_filter():
    global option
    option= str(calibration_option.get())
    if option=='Magnetometer sensor':
        
        try:
            for widget in extra.winfo_children():
                    widget.destroy() 
        except:
            pass
        try:
            for widget in extra1.winfo_children():
                    widget.destroy() 
        except:
            pass       
        check_butt = Button(master = mainw, command =lambda:run_magnetic(), height = 1, width =10, text = "Load data") 
        check_butt.grid(row=3,column=0)


        def run_magnetic():
            global filename
            import load as load
            data,numdata,empty,NaN=load.l()
            global extra

            #Import all Variables
            
            import newvar as allvar
            alldata,acc_sensor,ii,\
            sensor_name,encoder_finder,encoder_diff,pressure,voltage,depth,aa,\
            thetadict,angleall,degreedict,degreeall,ljoint,larm,lcont,radius,\
            radiusdict,radiusall,angle,degree,inclinationdirection,incline,\
            magnetic_declination,inclination_x,inclination_y,anglecaliball,degreecaliball,\
            radiuscaliball,thetadict_calib,degree_calibdict,radius_calibdict,\
            angle_calibrated,degree_calibrated,radius_calibrated\
            =allvar.allvariable(numdata,data,empty,NaN)
            ###############################################################################
            #Accelerometer sensors
            
            import newacc as newacc
            acc_sensor=newacc.acc(numdata,data,alldata,ii,sensor_name,acc_sensor,\
              NaN,encoder_finder,encoder_diff)
            ###############################################################################
            #MAGNETOMETER SESNOR
            
            import newmag as mag_sensor
            magnetic_x,magnetic_y,magnetic_z,magneticfield_x_uncalibrated,\
            magneticfield_y_uncalibrated,magneticfield_z_uncalibrated,absolute_uncalibrated_magneticfield\
            =mag_sensor.magnetometer_sensor(numdata,data,NaN,encoder_finder,encoder_diff)
            ##############################################################################
            #Uncalibrated rotation of the logger   
            
            import newcalc 
            sum_mag_declination,inclinationdirection\
            =newcalc.inclination(numdata,inclinationdirection,incline,\
            magnetic_declination,acc_sensor,magnetic_x,magnetic_y,magnetic_z)    
            ############################################################################### 
            #Inclination 
            
            import newinc
            inclination_x,inclination_y,xarm_calibrated,yarm_calibrated,zarm_calibrated,\
            xcalibrated,ycalibrated,zcalibrated,xx_calibrated,yy_calibrated,zz_calibrated,\
            inclinationdirection,incline,x3_avg,y3_avg,z3_avg,inclination_xaxis_avg,inclination_yaxis_avg\
            =newinc.calib(numdata,acc_sensor,sum_mag_declination,\
            inclination_x,inclination_y)
            ###############################################################################

            try:
                for widget in extra.winfo_children():
                        widget.destroy() 
            except:
                pass
            try:
                for widget in extra1.winfo_children():
                        widget.destroy() 
            except:
                pass      
            clean()
            extra = Frame(mainw )
            extra.grid(row=5,column=1, sticky='W')                  
  
            
            time=np.zeros(numdata)
            absolute_magneticfield=np.zeros(numdata)
            # Find the magnetic field by using tilt compensation
            #Find magnetic declination
            for i in range(numdata):
                time[i]=(0.5*i)
                tilt_compensation=([(magnetic_x[i]*math.cos(math.radians(y3_avg)))+(magnetic_y[i]*math.sin(math.radians(y3_avg))*math.sin(math.radians(x3_avg)))-(magnetic_z[i]*math.cos(math.radians(x3_avg))*math.sin(math.radians(x3_avg)))],[(magnetic_y[i]*math.cos(math.radians(y3_avg)))-(magnetic_z[i]*math.sin(math.radians(y3_avg)))],[0])
                tilt_compensation=np.array(tilt_compensation)
                absolute_magneticfield[i]=np.sqrt((tilt_compensation[0]**2)+(tilt_compensation[1]**2)+(magnetic_z[i]**2)) 
                magnetic_declination[i]=180*math.atan2(((tilt_compensation[1])),(tilt_compensation[0]))/math.pi %360
            diff_magnetic=np.diff(magnetic_declination)
            # absolute_magneticfield[i+1]=absolute_magneticfield[i]
            for i in range (numdata-1):
                if diff_magnetic[i]>300:
                    diff_magnetic[i]=diff_magnetic[i]-360
                if diff_magnetic[i]<-300:
                    diff_magnetic[i]=360+diff_magnetic[i]  
            sum_mag_declination=np.zeros(numdata-1)
            for i in range (numdata-1):
                sum_mag_declination[i]=np.sum(diff_magnetic[0:i+1]) 

            # Find magnetic offset in x and y axes
            magnetic_offset_x=(min(magneticfield_x_uncalibrated)+max(magneticfield_x_uncalibrated))/2
            magnetic_offset_y=(min(magneticfield_y_uncalibrated)+max(magneticfield_y_uncalibrated))/2
            disc = LabelFrame(mainw, text="Magnetic offset result:", font="Arial 12 bold italic")
            disc.grid(row=4,column=0)
            scrollbar = Scrollbar(disc)
            discrip = tk.Text(disc, height=5, width=70, yscrollcommand=scrollbar.set, font="Arial 14")
            scrollbar.config(command=discrip.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            discrip.pack(side="left")
            discrip.insert(END,'magnetic_x_offset= '+str(round(-magnetic_offset_x,3))+' μT'+'\nmagnetic_y_offset= '+str(round(-magnetic_offset_y,3))+' μT'+'\nRotation degree= '+str(round(sum_mag_declination[numdata-2],3))+' Deg'+'\nCalibrated absolute magnetic field= '+str(round(np.mean(absolute_magneticfield),3))+' μT'+'\nUncalibrated absolute magnetic field= '+str(round(np.mean(absolute_uncalibrated_magneticfield),3))+' μT')  
            #Visualize magnetic field
            plotplot = LabelFrame(mainw)
            plotplot.grid(row=5,column=0)
            mainw.update_idletasks()
            f=plt.figure()
            for widget in plotplot.winfo_children():
                widget.destroy()
            canvas = FigureCanvasTkAgg(f, master =plotplot)
            canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)
    
            toolbarFrame = Frame(master=plotplot)
            toolbarFrame.grid(row=1,column=0)
            toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)
    
            canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4) 
            for widget in plotplot.winfo_children():
                widget.destroy()        
            canvas = Canvas(master =plotplot,bg='white',width='950',height='600')
            canvas.grid(row=0,column=0, padx=4, pady=4)                             
            f = plt.figure()
            ax = f.add_subplot(111)
            plt.plot(time,magnetic_x,'r',label="mag_X_calibrated");plt.plot(time,magnetic_y,'b',label="mag_Y_calibrated");plt.plot(time,magnetic_z,'g',label="mag_Z_calibrated");plt.plot(time,absolute_magneticfield,'k',label="mag_abs__calibrated")
            plt.plot(time,magneticfield_x_uncalibrated,'r--',label="mag_X_uncalibrated");plt.plot(time,magneticfield_y_uncalibrated,'b--',label="mag_Y_uncalibrated");plt.plot(time,magneticfield_z_uncalibrated,'g--',label="mag_Z_uncalibrated");plt.plot(time,absolute_uncalibrated_magneticfield,'k--',label="mag_abs_uncalibrated")
            plt.legend(fontsize=7, ncol=2,loc='best')
            
            plt.xlabel('Time (sec)',fontsize=12)
            plt.ylabel('Magnetic field (μT)',fontsize=12)
            if np.mean(absolute_magneticfield)>70 or np.mean(absolute_magneticfield)<30:
                plt.title('WARNING!!!Magnetic field is out of the range\nCheck the logger')
            else:
                plt.title('Magnetometer field strength at -20°C',fontsize=12)
            
            if np.abs(sum_mag_declination[numdata-2])>340:
                    # try:
                    for widget in extra.winfo_children():
                        widget.destroy()
                    try:
                        for widget in extra1.winfo_children():
                                widget.destroy() 
                    except:
                        pass     
                    Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)                     
                    Label(extra,text="Rotation is around 360 DEG\nYou can use the offset\nto calibrate magnetometer\n---------------",font = "Arial 12 bold italic").grid(row=1,column=0)
                    Label(extra,text="Do you want to use this offset values\nas new magnetometer's offset? ",font = "Arial 12 bold italic").grid(row=2,column=0)
                    exy_mag = Button(master = extra, command =lambda:excel_mag(), height = 1, width =5, text = "YES") 
                    exy_mag.grid(row=3,column=0)
                    exn_mag = Button(master = extra, command =lambda:do_nothing(), height = 1, width =5, text = "NO") 
                    exn_mag.grid(row=3,column=1)
                    mainw.update_idletasks()
                    frame_canvas.config(width=screen_width-10,
                                        height=screen_height-10)
                    
                    # Set the canvas scrolling region
                    canvasf.config(scrollregion=canvasf.bbox("all"))   
                    #print the magnetic offsets in excel file
                    def excel_mag  ():
                        sheet.cell(row=2,column=1).value=np.abs(magnetic_offset_x)
                        sheet.cell(row=2,column=2).value=np.abs(magnetic_offset_y)
                        wb.save('calibration.xlsx')
                    def do_nothing():
                        pass

    
            else:
                for widget in extra.winfo_children():
                    widget.destroy()
                    try:
                        for widget in extra1.winfo_children():
                                widget.destroy() 
                    except:
                        pass    
                Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)                           
                Label(extra ,text="Rotation is NOT around 360 DEG\nDON'T use the offset\nto calibrate magnetometer\n---------------",font = "Arial 12 bold italic").grid(row=1,column=0)
                Label(extra ,text="Do you want to use this offset values\nas new magnetometer's offset? ",font = "Arial 12 bold italic").grid(row=2,column=0)
                exy_mag = Button(master = extra , command =lambda:excel_mag(), height = 1, width =5, text = "YES") 
                exy_mag.grid(row=3,column=0)
                exn_mag = Button(master = extra , command =lambda:do_nothing(), height = 1, width =5, text = "NO") 
                exn_mag.grid(row=3,column=1)
                mainw.update_idletasks()
                frame_canvas.config(width=screen_width-10,
                                    height=screen_height-10)
                
                # Set the canvas scrolling region
                canvasf.config(scrollregion=canvasf.bbox("all"))
                #print the magnetic offsets in excel file
                def excel_mag  ():
                    sheet.cell(row=2,column=1).value=np.abs(magnetic_offset_x)
                    sheet.cell(row=2,column=2).value=np.abs(magnetic_offset_y)
                    wb.save('calibration.xlsx')
                def do_nothing():
                    pass
            for widget in plotplot.winfo_children():
                widget.destroy()   
            canvas = FigureCanvasTkAgg(f, master =plotplot)
            canvas.draw()

            canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)

            toolbarFrame = Frame(master=plotplot)
            toolbarFrame.grid(row=1,column=0)
            toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)

            canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4)

################################GYROSCOPE CALIBRATION#########################
    if option=='Gyro':
        try:
            for widget in extra.winfo_children():
                    widget.destroy() 
        except:
            pass
        try:
            for widget in extra1.winfo_children():
                    widget.destroy() 
        except:
            pass  
        
        check_butt = Button(master = mainw, command =lambda:run_gyro(), height = 1, width =10, text = "Load data") 
        check_butt.grid(row=3,column=0)
        def run_gyro():
            global filename
            import load as load
            global extra
            data,numdata,empty,NaN=load.l()
            #Import all Variables
            
            import newvar as allvar
            alldata,acc_sensor,ii,\
            sensor_name,encoder_finder,encoder_diff,pressure,voltage,depth,aa,\
            thetadict,angleall,degreedict,degreeall,ljoint,larm,lcont,radius,\
            radiusdict,radiusall,angle,degree,inclinationdirection,incline,\
            magnetic_declination,inclination_x,inclination_y,anglecaliball,degreecaliball,\
            radiuscaliball,thetadict_calib,degree_calibdict,radius_calibdict,\
            angle_calibrated,degree_calibrated,radius_calibrated\
            =allvar.allvariable(numdata,data,empty,NaN)
            ##############################################################################
            #Accelerometer sensors
            
            import newacc as newacc
            acc_sensor=newacc.acc(numdata,data,alldata,ii,sensor_name,acc_sensor,\
              NaN,encoder_finder,encoder_diff)
            ##############################################################################
                
            #MAGNETOMETER SESNOR
            
            import newmag as mag_sensor
            magnetic_x,magnetic_y,magnetic_z,magneticfield_x_uncalibrated,\
            magneticfield_y_uncalibrated,magneticfield_z_uncalibrated,absolute_uncalibrated_magneticfield\
            =mag_sensor.magnetometer_sensor(numdata,data,NaN,encoder_finder,encoder_diff)
    
            ###############################################################################
            ##GYROSCOPE SENSOR
            
            import newgyro as gyro_sensor
            gyro_x,gyro_y,gyro_z,gyro_xuncalibrated,gyro_yuncalibrated,gyro_zuncalibrated,\
            gyro_xcalibrated,gyro_ycalibrated,gyro_zcalibrated,gyro_rotation_uncalibrated,gyro_rotation_calibrated,\
            gyro_rotation_xuncalibrated,gyro_rotation_xcalibrated,gyro_rotation_yuncalibrated,\
            gyro_rotation_ycalibrated\
            =gyro_sensor.gyroscope(numdata,data,NaN,encoder_finder,encoder_diff)
            ##############################################################################
            #Uncalibrated rotation of the logger   
            
            import newcalc 
            sum_mag_declination,inclinationdirection\
            =newcalc.inclination(numdata,inclinationdirection,incline,\
            magnetic_declination,acc_sensor,magnetic_x,magnetic_y,magnetic_z)    

            ########

            
            time=np.zeros(numdata)
            gyro_offset_uncalibrated=np.mean(gyro_zuncalibrated)
            gyro_rotation_uncalibrated=(np.sum(gyro_zuncalibrated))
            gyro_rotation_calibrated=(np.sum(gyro_zcalibrated))            
            gyro_time_uncalibrated=np.zeros(numdata)
            gyro_time_calibrated=np.zeros(numdata) 
            gyro_offset_xuncalibrated=np.mean(gyro_xuncalibrated)
            gyro_rotation_xuncalibrated=(np.sum(gyro_xuncalibrated))
            gyro_rotation_xcalibrated=(np.sum(gyro_xcalibrated))            
            gyro_time_xuncalibrated=np.zeros(numdata)
            gyro_time_xcalibrated=np.zeros(numdata)
            gyro_offset_yuncalibrated=np.mean(gyro_yuncalibrated)
            gyro_rotation_yuncalibrated=(np.sum(gyro_yuncalibrated))
            gyro_rotation_ycalibrated=(np.sum(gyro_ycalibrated))            
            gyro_time_yuncalibrated=np.zeros(numdata)
            gyro_time_ycalibrated=np.zeros(numdata)
            
            try:
                for widget in extra.winfo_children():
                        widget.destroy() 
            except:
                pass
            try:
                for widget in extra1.winfo_children():
                        widget.destroy() 
            except:
                pass    
            clean()                 
            extra = Frame(mainw)
            extra.grid(row=5,column=1, sticky='W')                
            for i in range(numdata):
                time[i]=(0.5*i)
                gyro_time_uncalibrated[i]=np.sum(gyro_zuncalibrated[0:i])
                gyro_time_calibrated[i]=np.sum(gyro_zcalibrated[0:i])
                gyro_time_xuncalibrated[i]=np.sum(gyro_xuncalibrated[0:i])
                gyro_time_xcalibrated[i]=np.sum(gyro_xcalibrated[0:i])
                gyro_time_yuncalibrated[i]=np.sum(gyro_yuncalibrated[0:i])
                gyro_time_ycalibrated[i]=np.sum(gyro_ycalibrated[0:i])
            disc = LabelFrame(mainw, text="Gyroscope offset result:", font="Arial 12 bold italic")
            disc.grid(row=4,column=0)
            scrollbar = Scrollbar(disc)
            discrip = tk.Text(disc, height=5, width=70, yscrollcommand=scrollbar.set, font="Arial 14")
            scrollbar.config(command=discrip.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            discrip.pack(side="left")
            discrip.insert(END,'gyro_z_offset= '+str(round(-gyro_offset_uncalibrated,3))+' Deg'+'\nGyro rotation uncalibrated= '+str(round(gyro_rotation_uncalibrated,3))+' Deg'+'\nGyro rotation calibrated= '+str(round(gyro_rotation_calibrated,3))+' Deg\n'\
                           'gyro_x_offset= '+str(round(-gyro_offset_xuncalibrated,3))+' Deg'+'\nGyro x uncalibrated= '+str(round(gyro_rotation_xuncalibrated,3))+' Deg'+'\nGyro x calibrated= '+str(round(gyro_rotation_xcalibrated,3))+' Deg\n'\
                               'gyro_y_offset= '+str(round(-gyro_offset_yuncalibrated,3))+' Deg'+'\nGyro y uncalibrated= '+str(round(gyro_rotation_yuncalibrated,3))+' Deg'+'\nGyro y calibrated= '+str(round(gyro_rotation_ycalibrated,3))+' Deg')  
            #Visualize the gyroscope sensor
            plotplot = LabelFrame(mainw)
            plotplot.grid(row=5,column=0)
            f=plt.figure()
            for widget in plotplot.winfo_children():
                widget.destroy()
            canvas = FigureCanvasTkAgg(f, master =plotplot)
            canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)
    
            toolbarFrame = Frame(master=plotplot)
            toolbarFrame.grid(row=1,column=0)
            toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)
    
            canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4) 
            for widget in plotplot.winfo_children():
                widget.destroy()        
            canvas = Canvas(master =plotplot,bg='white',width='900',height='500')
            canvas.grid(row=0,column=0, padx=4, pady=4)                             
            f = plt.figure()
            ax = f.add_subplot(111)
            plt.plot(time,gyro_time_xuncalibrated,'r--',label="Gyro_rotation X_axis uncalibrated");plt.plot(time,gyro_time_yuncalibrated,'b--',label="Gyro_rotation Y_axis uncalibrated");plt.plot(time,gyro_time_uncalibrated,'g--',label="Gyro_rotation Z_axis uncalibrated");
            plt.plot(time,gyro_time_xcalibrated,'r',label="Gyro_rotation X_axis calibrated");plt.plot(time,gyro_time_ycalibrated,'b',label="Gyro_rotation Y_axis calibrated");plt.plot(time,gyro_time_calibrated,'g',label="Gyro_rotation Z_axis calibrated")
            plt.legend(fontsize=9, loc='best')
            plt.xlabel('Time (sec)',fontsize=12)
            plt.ylabel('Rotation (Deg)',fontsize=12)
            plt.title('Gyroscope rotation before and after calibration at -20°C',fontsize=12)
            if np.abs(sum_mag_declination[numdata-2])<7:
                for widget in extra.winfo_children():
                    widget.destroy()  
                Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)                               
                Label(extra,text="Logger almost fixed\n---------------\n",font = "Arial 12 bold italic").grid(row=1,column=0)
                Label(extra,text="Do you want to use this offset values\nas new gyroscope's offset?\n \n",font = "Arial 12 bold italic").grid(row=2,column=0)
                exy_gyro = Button(master = extra, command =lambda:excel_gyro(), height = 1, width =5, text = "YES") 
                exy_gyro.grid(row=3,column=0)
                exn_gyro = Button(master = extra, command =lambda:do_nothing(), height = 1, width =5, text = "NO") 
                exn_gyro.grid(row=3,column=1)
                mainw.update_idletasks()
                frame_canvas.config(width=screen_width-10,
                                    height=screen_height-10)
                
                # Set the canvas scrolling region
                canvasf.config(scrollregion=canvasf.bbox("all")) 
                #print the gyro offsets in excel file
                def excel_gyro  ():
                    sheet.cell(row=2,column=3).value=-gyro_offset_uncalibrated
                    sheet.cell(row=2,column=4).value=-gyro_offset_xuncalibrated
                    sheet.cell(row=2,column=5).value=-gyro_offset_yuncalibrated
                    wb.save('calibration.xlsx')
                def do_nothing():
                    pass
    
    
            else:
                for widget in extra.winfo_children():
                    widget.destroy()  
                try:
                    for widget in extra1.winfo_children():
                            widget.destroy() 
                except:
                    pass 
                Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)                            
                Label(extra,text="Logger is not fixed\nDON'T USE this offset as gyroscope offset\n---------------",font = "Arial 12 bold italic").grid(row=1,column=0)
                Label(extra,text="Do you want to use this offset values\nas new gyroscope's offset?\n \n",font = "Arial 12 bold italic").grid(row=2,column=0)
                exy_gyro = Button(master = extra, command =lambda:excel_gyro(), height = 1, width =5, text = "YES") 
                exy_gyro.grid(row=3,column=0)
                exn_gyro = Button(master = extra, command =lambda:do_nothing(), height = 1, width =5, text = "NO") 
                exn_gyro.grid(row=3,column=1)
                mainw.update_idletasks()
                frame_canvas.config(width=screen_width-10,
                                    height=screen_height-10)
                
                # Set the canvas scrolling region
                canvasf.config(scrollregion=canvasf.bbox("all"))  
                #print the gyro offsets in excel file
                def excel_gyro  ():
                    sheet.cell(row=2,column=3).value=-gyro_offset_uncalibrated
                    wb.save('calibration.xlsx')
                def do_nothing():
                    pass
            for widget in plotplot.winfo_children():
                widget.destroy()   
            canvas = FigureCanvasTkAgg(f, master =plotplot)
            canvas.draw()

            canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)

            toolbarFrame = Frame(master=plotplot)
            toolbarFrame.grid(row=1,column=0)
            toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)

            canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4)    

################################INCLINATION ANGLE CALIBRATION##################
    if option=='Inclination angle':
        try:
            for widget in extra.winfo_children():
                    widget.destroy() 
        except:
            pass
        try:
            for widget in extra1.winfo_children():
                    widget.destroy() 
        except:
            pass  
                
        check_butt = Button(master = mainw, command =lambda:run_inclination(), height = 1, width =10, text = "Load data") 
        check_butt.grid(row=3,column=0)
        def run_inclination():
            global filename
            import load as load
            global extra
            data,numdata,empty,NaN=load.l()
            #Import all Variables
            
            import newvar as allvar
            alldata,acc_sensor,ii,\
            sensor_name,encoder_finder,encoder_diff,pressure,voltage,depth,aa,\
            thetadict,angleall,degreedict,degreeall,ljoint,larm,lcont,radius,\
            radiusdict,radiusall,angle,degree,inclinationdirection,incline,\
            magnetic_declination,inclination_x,inclination_y,anglecaliball,degreecaliball,\
            radiuscaliball,thetadict_calib,degree_calibdict,radius_calibdict,\
            angle_calibrated,degree_calibrated,radius_calibrated\
            =allvar.allvariable(numdata,data,empty,NaN)
            ###############################################################################
            #Accelerometer sensors
            
            import newacc as newacc
            acc_sensor=newacc.acc(numdata,data,alldata,ii,sensor_name,acc_sensor,\
              NaN,encoder_finder,encoder_diff)
            ###############################################################################
            #MAGNETOMETER SESNOR
            
            import newmag as mag_sensor
            magnetic_x,magnetic_y,magnetic_z,magneticfield_x_uncalibrated,\
            magneticfield_y_uncalibrated,magneticfield_z_uncalibrated,absolute_uncalibrated_magneticfield\
            =mag_sensor.magnetometer_sensor(numdata,data,NaN,encoder_finder,encoder_diff)
            ##############################################################################
            ##GYROSCOPE SENSOR
            
            import newgyro as gyro_sensor
            gyro_x,gyro_y,gyro_z,gyro_xuncalibrated,gyro_yuncalibrated,gyro_zuncalibrated,\
            gyro_xcalibrated,gyro_ycalibrated,gyro_zcalibrated,gyro_rotation_uncalibrated,gyro_rotation_calibrated,\
            gyro_rotation_xuncalibrated,gyro_rotation_xcalibrated,gyro_rotation_yuncalibrated,\
            gyro_rotation_ycalibrated\
            =gyro_sensor.gyroscope(numdata,data,NaN,encoder_finder,encoder_diff)
            ###############################################################################   
            #Uncalibrated rotation of the logger   
            
            import newcalc 
            sum_mag_declination,inclinationdirection\
            =newcalc.inclination(numdata,inclinationdirection,incline,\
            magnetic_declination,acc_sensor,magnetic_x,magnetic_y,magnetic_z)    
            ############################################################################### 
            #Inclination 
            
            import newinc
            inclination_x,inclination_y,xarm_calibrated,yarm_calibrated,zarm_calibrated,\
            xcalibrated,ycalibrated,zcalibrated,xx_calibrated,yy_calibrated,zz_calibrated,\
            inclinationdirection,incline,x3_avg,y3_avg,z3_avg,inclination_xaxis_avg,inclination_yaxis_avg\
            =newinc.calib(numdata,acc_sensor,sum_mag_declination,\
            inclination_x,inclination_y)

            try:
                for widget in extra.winfo_children():
                        widget.destroy() 
            except:
                pass
            try:
                for widget in extra1.winfo_children():
                        widget.destroy() 
            except:
                pass   
            clean()                  
            extra = Frame(mainw)
            extra.grid(row=5,column=1, sticky='W')                      
            inclinationdirection=np.zeros(numdata)
            for i in range (numdata):
                    if math.isnan(inclinationdirection[i])==True:
                        inclinationdirection[i]=inclinationdirection[i-1]
                    incline[i]=np.arccos(((y3_avg[i]))/(math.sqrt((y3_avg[i]**2)+(x3_avg[i]**2)+(z3_avg[i]**2))))
            inclinationdirection=(inclinationdirection*180)/math.pi %360
            incline=180-(incline*180)/math.pi 
            for i in range (numdata-1):  
        
                if math.isnan(incline[i])==True:
                     incline[i]=incline[i-1]                
                           
            disc = LabelFrame(mainw, text="Inclination result:", font="Arial 12 bold italic")
            disc.grid(row=4,column=0)
            scrollbar = Scrollbar(disc)
            discrip = tk.Text(disc, height=5, width=70, yscrollcommand=scrollbar.set, font="Arial 14")
            scrollbar.config(command=discrip.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            discrip.pack(side="left")
            discrip.insert(END,'Offset of inclination along_x_13= '+str(round(np.mean(inclination_x['inclination_xaxis_13']),3))+' Deg'+'\nOffset of inclination along_x_23= '+str(round(np.mean(inclination_x['inclination_xaxis_23']),3))+' Deg'+'\nOffset of inclination along_x_33= '+str(round(np.mean(inclination_x['inclination_xaxis_33']),3))+' Deg'+'\nOffset of inclination along_x_43= '+str(round(np.mean(inclination_x['inclination_xaxis_43']),3))+' Deg'+'\nOffset of inclination along_x_53= '+str(round(np.mean(inclination_x['inclination_xaxis_53']),3))+' Deg'+'\nOffset of inclination along_x_63= '+str(round(np.mean(inclination_x['inclination_xaxis_63']),3))+' Deg'+'\nOffset of inclination along_x_73= '+str(round(np.mean(inclination_x['inclination_xaxis_73']),3))+' Deg'+'\nOffset of inclination along_x_83= '+str(round(np.mean(inclination_x['inclination_xaxis_83']),3))+' Deg'+'\nOffset of inclination along_y_13= '+str(round(np.mean(inclination_y['inclination_yaxis_13']),3))+' Deg'+'\nOffset of inclination along_y_23= '+str(round(np.mean(inclination_y['inclination_yaxis_23']),3))+' Deg'+'\nOffset of inclination along_y_33= '+str(round(np.mean(inclination_y['inclination_yaxis_33']),3))+' Deg'+'\nOffset of inclination along_y_43= '+str(round(np.mean(inclination_y['inclination_yaxis_43']),3))+' Deg'+'\nOffset of inclination along_y_53= '+str(round(np.mean(inclination_y['inclination_yaxis_53']),3))+' Deg'+'\nOffset of inclination along_y_63= '+str(round(np.mean(inclination_y['inclination_yaxis_63']),3))+' Deg'+'\nOffset of inclination along_y_73= '+str(round(np.mean(inclination_y['inclination_yaxis_73']),3))+' Deg'+'\nOffset of inclination along_y_83= '+str(round(np.mean(inclination_y['inclination_yaxis_83']),3))+' Deg')      
            plotplot = LabelFrame(mainw)
            plotplot.grid(row=5,column=0)
            time=np.zeros(numdata)
            for i in range(numdata):
                time[i]=(0.5*i)
                inclination_yaxis_avg[numdata-1]=inclination_yaxis_avg[numdata-2]
                inclination_xaxis_avg[numdata-1]=inclination_xaxis_avg[numdata-2]
                incline[numdata-1]=incline[numdata-2]

            incline[numdata-1]=incline[numdata-2]                
            f=plt.figure()
            for widget in plotplot.winfo_children():
                widget.destroy()
            canvas = FigureCanvasTkAgg(f, master =plotplot)
            canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)
    
            toolbarFrame = Frame(master=plotplot)
            toolbarFrame.grid(row=1,column=0)
            toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)
    
            canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4) 
            for widget in plotplot.winfo_children():
                widget.destroy()        
            canvas = Canvas(master =plotplot,bg='white',width='900',height='500')
            canvas.grid(row=0,column=0, padx=4, pady=4)              
            #print the magnetic offsets in excel file               
            f = plt.figure()
            ax = f.add_subplot(111)
            plt.plot(time[0:numdata],inclination_xaxis_avg,'r',label="Average inclination_along_x");plt.plot(time[0:numdata],inclination_yaxis_avg,'b',label="Average inclination_along_y")
            
            plt.xlabel('Time (sec)')
            plt.ylabel('Inclination (Deg)')   
            plt.legend(fontsize=7,loc='best')
            plt.title('Average inclination')
            global abs_inc
            abs_inc=(np.mean(incline))
            abs_inc=round(abs_inc,2)
            global x_inc
            global y_inc
            plt.title('Calibrated absolute inclination='+str(abs_inc)+'deg')
            Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)
            Label(extra,text="Measured inclination along x_axis ( + and - is important)",font = "Arial 10 bold italic").grid(row=1,column=0)
            x_inc = Entry(master=extra,width=5)
            x_inc.grid(row=2,column=0)

            Label(extra,text="Measured inclination along y_axis ( + and - is important)",font = "Arial 10 bold italic").grid(row=3,column=0)
            y_inc = Entry(master=extra,width=5)
            y_inc.grid(row=4,column=0)
            exy_gyro = Button(master = extra, command =lambda:confirm_inc(), height = 2, width =10, text = "CONFIRM") 
            exy_gyro.grid(row=5,column=0)
            mainw.update_idletasks()
            frame_canvas.config(width=screen_width-10,
                                height=screen_height-10)
            
            # Set the canvas scrolling region
            canvasf.config(scrollregion=canvasf.bbox("all"))            
            def confirm_inc():
                global x_inc
                global y_inc
                global extra
                global extra1
                x_incline=(float(x_inc.get()))
                y_incline=(float(y_inc.get()))                    
                for widget in extra.winfo_children():
                        widget.destroy()
                extra1 =Frame(mainw)
                extra1.grid(row=5,column=1, sticky='W')                       
                if np.abs(x_incline)<0.1 and np.abs(y_incline)<0.1:
                    Label(extra1,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)
                    Label(extra1,text="Measured inclination along x_axis ( + and - is important)",font = "Arial 10 bold italic").grid(row=1,column=0)
                    x_inc = Entry(master=extra1,width=5)
                    x_inc.grid(row=2,column=0)
                    x_inc.insert(0,x_incline)
                    x_inc=float(x_inc.get())
                    Label(extra1,text="Measured inclination along y_axis ( + and - is important)",font = "Arial 10 bold italic").grid(row=3,column=0)
                    y_inc = Entry(master=extra1,width=5)
                    y_inc.grid(row=4,column=0)
                    y_inc.insert(0,y_incline)
                    y_inc=float(y_inc.get())
                    exy_gyro = Button(master = extra1, command =lambda:confirm_inc(), height = 2, width =10, text = "CONFIRM") 
                    exy_gyro.grid(row=5,column=0)

                    Label(extra1,text="Logger is almost vertical\n if you want to \ncalibrate the x and y_axis of accelerometer\npress *Confirm logger is vertical*\n",font = "Arial 10 bold italic").grid(row=6,column=0)
                    verty_inc = Button(master = extra1, command =lambda:confirm_vert(), height = 3, width =20, text = "Confirm logger is vertical\ncalibrate x and y_axis") 
                    verty_inc.grid(row=7,column=0)
                    vertn_inc = Button(master = extra1, command =lambda:confirm_nvert(), height = 3, width =20, text = "Logger is not vertical\ncalibrate z_axis") 
                    vertn_inc.grid(row=8,column=0)
                    mainw.update_idletasks()
                    frame_canvas.config(width=screen_width-10,
                                        height=screen_height-10)
                    
                    # Set the canvas scrolling region
                    canvasf.config(scrollregion=canvasf.bbox("all"))  
                    #print the reference accelerometer offsets in excel file
                    def confirm_vert():
                        for k in range (8):
                            k=k+1
                            sheet.cell(row=k+1,column=7).value=-np.mean(inclination_x['inclination_xaxis_'+str(k)+'3'][0:numdata-1])
                            sheet.cell(row=k+1,column=8).value=-np.mean(inclination_y['inclination_yaxis_'+str(k)+'3'][0:numdata-1])
                        wb.save('calibration.xlsx')
                else:   
                    for widget in extra.winfo_children():
                            widget.destroy()
                    extra1 = Frame(mainw)
                    extra1.grid(row=5,column=1, sticky='W')  
                    Label(extra1,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)                     
                    Label(extra1,text="Measured inclination along x_axis ( + and - is important)",font = "Arial 10 bold italic").grid(row=1,column=0)
                    x_inc = Entry(master=extra1,width=5)
                    x_inc.grid(row=2,column=0)
                    x_inc.insert(0,x_incline)
                    x_inc=float(x_inc.get())
                    Label(extra1,text="Measured inclination along y_axis ( + and - is important)",font = "Arial 10 bold italic").grid(row=3,column=0)
                    y_inc = Entry(master=extra1,width=5)
                    y_inc.grid(row=4,column=0)
                    y_inc.insert(0,y_incline)
                    y_inc=float(y_inc.get())
                    exy_gyro = Button(master = extra1, command =lambda:confirm_inc(), height = 2, width =10, text = "CONFIRM") 
                    exy_gyro.grid(row=5,column=0)

                    Label(extra1,text="Logger is not vertical if you want to \ncalibrate the x and y_axis of\naccelerometer make sure that logger is vertical\nif you want calibrate z_axis press\n*Logger is not vertical*",font = "Arial 10 bold italic").grid(row=6,column=0)
                    verty_inc = Button(master = extra1, command =lambda:confirm_vert(), height = 3, width =20, text = "Confirm logger is vertical\ncalibrate x and y_axis") 
                    verty_inc.grid(row=7,column=0)
                    vertn_inc = Button(master = extra1, command =lambda:confirm_nvert(), height = 3, width =20, text = "Logger is not vertical\ncalibrate z_axis") 
                    vertn_inc.grid(row=8,column=0) 
                    mainw.update_idletasks()
                    frame_canvas.config(width=screen_width-10,
                                        height=screen_height-10)
                    
                    # Set the canvas scrolling region
                    canvasf.config(scrollregion=canvasf.bbox("all"))   
                    #print the reference accelerometer offsets in excel file
                    def confirm_vert():
                        for k in range (8):
                            k=k+1
                            sheet.cell(row=k+1,column=7).value=-np.mean(inclination_x['inclination_xaxis_'+str(k)+'3'][0:numdata-1])
                            sheet.cell(row=k+1,column=8).value=-np.mean(inclination_y['inclination_yaxis_'+str(k)+'3'][0:numdata-1])
                    wb.save('calibration.xlsx')
                    
                    def confirm_nvert(): 
                        from scipy.optimize import fsolve
                        for k in range (4):
                            k=k+1
                            def func(x):
                                 return [((180*math.atan((-x[0]/((x[1]**2)+(np.mean(ycalibrated[str(k)+'3'][0:numdata-1])**2))**(0.5)))/math.pi)-x_inc),
                                         ((180*math.atan((x[1]/((x[0]**2)+(np.mean(ycalibrated[str(k)+'3'][0:numdata-1])**2))**(0.5)))/math.pi)-y_inc),
                                         (((np.mean(xcalibrated[str(k)+'3'][0:numdata-1]))*np.cos(x[2]))+((np.mean(zcalibrated[str(k)+'3'][0:numdata-1]))*np.sin(x[2])) - x[1]),
                                         (((np.mean(xx_calibrated[str(k)+'3']))*np.sin(-x[2]))+((np.mean(zcalibrated[str(k)+'3'][0:numdata-1]))*np.cos(x[2])) - x[0])]
                            root = fsolve(func, [0,0,0,0])
                            sheet.cell(row=k+1,column=9).value=180*root[2]/math.pi

                        for k in range (6):
                            k=k+1
                            if k<5:
                                pass
                            else:
                                def func(x):
                                     return [((180*math.atan((-x[0]/((x[1]**2)+(np.mean(ycalibrated[str(k)+'3'][0:numdata-1])**2))**(0.5)))/math.pi)-x_inc),
                                             ((180*math.atan((x[1]/((x[0]**2)+(np.mean(ycalibrated[str(k)+'3'][0:numdata-1])**2))**(0.5)))/math.pi)-y_inc),
                                             (((np.mean(xcalibrated[str(k)+'3'][0:numdata-1]))*np.cos(x[2]))+((np.mean(zcalibrated[str(k)+'3'][0:numdata-1]))*np.sin(x[2])) - x[1]),
                                             (((np.mean(xx_calibrated[str(k)+'3']))*np.sin(-x[2]))+((np.mean(zcalibrated[str(k)+'3'][0:numdata-1]))*np.cos(x[2])) - x[0])]
                                root = fsolve(func, [0,0,0,0])
                                if 180*root[2]/math.pi<0:
                                    root[2]=180*root[2]/math.pi %180
                                    sheet.cell(row=k+1,column=9).value=root[2]
                                else:
                                    sheet.cell(row=k+1,column=9).value=180*root[2]/math.pi

                        for k in range (8):
                            k=k+1
                            if k<7:
                                pass
                            else:
                                def func(x):
                                     return [((180*math.atan((-x[0]/((x[1]**2)+(np.mean(ycalibrated[str(k)+'3'][0:numdata-1])**2))**(0.5)))/math.pi)-x_inc),
                                             ((180*math.atan((x[1]/((x[0]**2)+(np.mean(ycalibrated[str(k)+'3'][0:numdata-1])**2))**(0.5)))/math.pi)-y_inc),
                                             (((np.mean(xcalibrated[str(k)+'3'][0:numdata-1]))*np.cos(x[2]))+((np.mean(zcalibrated[str(k)+'3'][0:numdata-1]))*np.sin(x[2])) - x[1]),
                                             (((np.mean(xx_calibrated[str(k)+'3']))*np.sin(-x[2]))+((np.mean(zcalibrated[str(k)+'3'][0:numdata-1]))*np.cos(x[2])) - x[0])]
                                root = fsolve(func, [0,0,0,0])
                                if 180*root[2]/math.pi<0:
                                    root[2]=180*root[2]/math.pi %-180
                                    sheet.cell(row=k+1,column=9).value=root[2]
                                else:
                                    sheet.cell(row=k+1,column=9).value=180*root[2]/math.pi                            
                        
                        wb.save('calibration.xlsx')
     
            for widget in plotplot.winfo_children():
                widget.destroy()   
            canvas = FigureCanvasTkAgg(f, master =plotplot)
            canvas.draw()

            canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)

            toolbarFrame = Frame(master=plotplot)
            toolbarFrame.grid(row=1,column=0)
            toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)

            canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4)    

################################ARM ANGLE STEP1################################
    if option=='Arm radius step1':
        try:
            for widget in extra.winfo_children():
                    widget.destroy() 
        except:
            pass
        try:
            for widget in extra1.winfo_children():
                    widget.destroy() 
        except:
            pass  
                
        check_butt = Button(master = mainw, command =lambda:run_arm(), height = 1, width =10, text = "Load data") 
        check_butt.grid(row=3,column=0)
        def run_arm():
            global filename
            global numdata
            global num1
            try:
                num1=numdata
            except:
                pass
            ##Load Data

            def l():
                global countnew
                import decodea as decodea        
                path=os.path.abspath('calibration.txt')
                if platform.system()=='Darwin':
                    path=path.replace("/calibration.txt", "")
                if platform.system()=='Linux':
                    path=path.replace("/calibration.txt", "")    
                if platform.system()=='Windows':
                    path=path.replace("\calibration.txt", "")
                count=0    
                try:
                    if countnew==count:
                        try:    
                            pass # one file at a time    
                        except:
                            pass
                except:     
                    try: 
                        countnew=count
                        os.remove("test.raw") # one file at a time  
                        countrem=1
                    except:
                            pass
  
                Tk().withdraw()
                rawdata=filedialog.askopenfilename(filetypes=(("text file", "*.txt"), ("All files", "*.*")))
                raw=decodea.main(rawdata)
                fileObject = open("test.raw", encoding="utf8", errors='ignore')
                text_trans = fileObject.read()
                data=text_trans
                data=data.split ("\n")
                data=data[1:len(data)]
                numdata=len(data)-1
            
                data=str(data)
                data = data.replace("'", "")
                data = data.replace('"', "")
                data = data.replace(",", ";")
                data = data.replace(" ", "")
                data = data.replace(",", ";")
                data = data.replace(" ", "")
                data = data.replace(":", ";")
                data=data.split (";")
                data=' '.join(data).split()
                NaN = ("NaN")
                empty=[i for i,val in enumerate(data) if val=='[{}']
                
                return data,numdata,empty,NaN  

            data,numdata,empty,NaN=l()        
            ##############################################################################
            #Import all Variables
            
            import newvar as allvar
            alldata,acc_sensor,ii,\
            sensor_name,encoder_finder,encoder_diff,pressure,voltage,depth,aa,\
            thetadict,angleall,degreedict,degreeall,ljoint,larm,lcont,radius,\
            radiusdict,radiusall,angle,degree,inclinationdirection,incline,\
            magnetic_declination,inclination_x,inclination_y,anglecaliball,degreecaliball,\
            radiuscaliball,thetadict_calib,degree_calibdict,radius_calibdict,\
            angle_calibrated,degree_calibrated,radius_calibrated\
            =allvar.allvariable(numdata,data,empty,NaN)
            ###############################################################################
            #Accelerometer sensors
            
            import newacc as newacc
            acc_sensor=newacc.acc(numdata,data,alldata,ii,sensor_name,acc_sensor,\
              NaN,encoder_finder,encoder_diff)
            ###############################################################################
            #MAGNETOMETER SESNOR
            
            import newmag as mag_sensor
            magnetic_x,magnetic_y,magnetic_z,magneticfield_x_uncalibrated,\
            magneticfield_y_uncalibrated,magneticfield_z_uncalibrated,absolute_uncalibrated_magneticfield\
            =mag_sensor.magnetometer_sensor(numdata,data,NaN,encoder_finder,encoder_diff)
            ##############################################################################
            ##GYROSCOPE SENSOR
            
            import newgyro as gyro_sensor
            gyro_x,gyro_y,gyro_z,gyro_xuncalibrated,gyro_yuncalibrated,gyro_zuncalibrated,\
            gyro_xcalibrated,gyro_ycalibrated,gyro_zcalibrated,gyro_rotation_uncalibrated,gyro_rotation_calibrated,\
            gyro_rotation_xuncalibrated,gyro_rotation_xcalibrated,gyro_rotation_yuncalibrated,\
            gyro_rotation_ycalibrated\
            =gyro_sensor.gyroscope(numdata,data,NaN,encoder_finder,encoder_diff)
            ###############################################################################
            #Calculate the uncalibrated angle and radius of the arms
            
            import newangleradius   
            degreeall,radiusall\
            =newangleradius.ang_rad(numdata, angle, degree, aa, thetadict,\
            angleall, degreedict, degreeall, ljoint, larm, lcont, radius,\
            radiusdict, radiusall,acc_sensor)        
            ###############################################################################   
            #Uncalibrated rotation of the logger   
            
            import newcalc 
            sum_mag_declination,inclinationdirection\
            =newcalc.inclination(numdata,inclinationdirection,incline,\
            magnetic_declination,acc_sensor,magnetic_x,magnetic_y,magnetic_z)    
            ############################################################################### 
            #Inclination 
            
            import newinc
            inclination_x,inclination_y,xarm_calibrated,yarm_calibrated,zarm_calibrated,\
            xcalibrated,ycalibrated,zcalibrated,xx_calibrated,yy_calibrated,zz_calibrated,\
            inclinationdirection,incline,x3_avg,y3_avg,z3_avg,inclination_xaxis_avg,inclination_yaxis_avg\
            =newinc.calib(numdata,acc_sensor,sum_mag_declination,\
            inclination_x,inclination_y)
            ###############################################################################
            ##ARM CALIBRATION
            
            import newarmcalib
            acc_sensor,radius_calibdict,degree_calibdict\
            =newarmcalib.arm(acc_sensor,xarm_calibrated,yarm_calibrated,zarm_calibrated,\
            numdata,anglecaliball,degreecaliball,radiuscaliball,thetadict_calib,degree_calibdict,radius_calibdict,\
            angle_calibrated,degree_calibrated,radius_calibrated,aa,lcont,larm,ljoint) 
                
            global extra
            
            try:
                for widget in extra.winfo_children():
                        widget.destroy()
            except:
                pass
            clean()
            extra = Frame(mainw)
            extra.grid(row=5,column=1, sticky='W')                  
            disc = LabelFrame(mainw, text="Arm radius:", font="Arial 12 bold italic")
            disc.grid(row=4,column=0)
            scrollbar = Scrollbar(disc)
            discrip = tk.Text(disc, height=5, width=70, yscrollcommand=scrollbar.set, font="Arial 14")
            scrollbar.config(command=discrip.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            discrip.pack(side="left")
            discrip.insert(END,'degree arm_11= '+str(round(np.mean(degree_calibdict['degreecalib_11_13'][0:numdata-1]),2))+'  deg'+'\ndegree arm_21= '+str(round(np.mean(degree_calibdict['degreecalib_21_23'][0:numdata-1]),2))+'  deg'+'\ndegree arm_31= '+str(round(np.mean(degree_calibdict['degreecalib_31_33'][0:numdata-1]),2))+'  deg'+'\ndegree arm_41= '+str(round(np.mean(degree_calibdict['degreecalib_41_43'][0:numdata-1]),2))+'  deg'+'\ndegree arm_51= '+str(round(np.mean(degree_calibdict['degreecalib_51_53'][0:numdata-1]),2))+'  deg'+'\ndegree arm_61= '+str(round(np.mean(degree_calibdict['degreecalib_61_63'][0:numdata-1]),2))+'  deg'+'\ndegree arm_71= '+str(round(np.mean(degree_calibdict['degreecalib_71_73'][0:numdata-1]),2))+'  deg'+'\ndegree arm_81= '+str(round(np.mean(degree_calibdict['degreecalib_81_83'][0:numdata-1]),2))+'  deg'+'\ndegree arm_12= '+str(round(np.mean(degree_calibdict['degreecalib_12_13'][0:numdata-1]),2))+'  deg'+'\ndegree arm_22= '+str(round(np.mean(degree_calibdict['degreecalib_22_23'][0:numdata-1]),2))+'  deg'+'\ndegree arm_32= '+str(round(np.mean(degree_calibdict['degreecalib_32_33'][0:numdata-1]),2))+'  deg'+'\ndegree arm_42= '+str(round(np.mean(degree_calibdict['degreecalib_42_43'][0:numdata-1]),2))+'  deg'+'\ndegree arm_52= '+str(round(np.mean(degree_calibdict['degreecalib_52_53'][0:numdata-1]),2))+'  deg'+'\ndegree arm_62= '+str(round(np.mean(degree_calibdict['degreecalib_62_63'][0:numdata-1]),2))+'  deg'+'\ndegree arm_72= '+str(round(np.mean(degree_calibdict['degreecalib_72_73'][0:numdata-1]),2))+'  deg'+'\ndegree arm_82= '+str(round(np.mean(degree_calibdict['degreecalib_82_83'][0:numdata-1]),2))+'  deg')      
            plotplot = LabelFrame(mainw)
            plotplot.grid(row=5,column=0)
            time=np.zeros(numdata)
            Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)
            Label(extra,text="Choose arm number ",font = "Arial 10 bold italic").grid(row=1,column=0)
            global arm_selector 
            global measure_selector 
            arm_selector = StringVar(extra)
            arm_selector.set(' Arm number ')
            filterop = OptionMenu(extra,arm_selector, *sorted({'11','12','21','22','31','32','41','42','51','52','61','62','71','72','81','82'}),)
            filterop.grid(row=2, column=0)
            measure_selector = StringVar(extra)
            measure_selector.set('     Measurement holes     ')
            measureop = OptionMenu(extra,measure_selector, *sorted({'third hole,second hole','third hole,first hole','second hole,first hole'}),)
            measureop.grid(row=3, column=0)
            
            verty_inc = Button(master = extra, command =lambda:confirm_arm(), height = 2, width =10, text = "CONFIRM") 
            verty_inc.grid(row=4,column=0)
            mainw.update_idletasks()
            frame_canvas.config(width=screen_width-10,
                                height=screen_height-10)
            
            # Set the canvas scrolling region
            canvasf.config(scrollregion=canvasf.bbox("all"))            
            def confirm_arm():
                global arm_selector
                global measure_selector 
                global arm_num
                global measure_hole
                global measure_mm
                global arm_rad
                global extra
              
                for widget in extra.winfo_children():
                    widget.destroy()
                extra = Frame(mainw)
                extra.grid(row=5,column=1, sticky='W')    
                Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)                   
                Label(extra,text="Choose arm number",font = "Arial 10 bold italic").grid(row=1,column=0)
                global arm_selector 

                filterop = OptionMenu(extra,arm_selector, *sorted({'11','12','21','22','31','32','41','42','51','52','61','62','71','72','81','82'}),)
                filterop.grid(row=2, column=0)
                measureop = OptionMenu(extra,measure_selector, *sorted({'third hole,second hole','third hole,first hole','second hole,first hole'}),)
                measureop.grid(row=3, column=0)                
                verty_inc = Button(master = extra, command =lambda:confirm_arm(), height = 2, width =10, text = "CONFIRM") 
                verty_inc.grid(row=4,column=0)

                arm_num=str(arm_selector.get())
                measure_hole=str(measure_selector.get())
                if measure_hole=='third hole,second hole':
                    measure_mm=str(6.95)
                if measure_hole=='third hole,first hole':
                    measure_mm=str(13.68)
                if measure_hole=='second hole,first hole':
                    measure_mm=str(6.73)
                for k in range (8):
                    k=k+1
                    if arm_num==str(k)+'1':    
                        lab=Label(extra,text='Guess the offset to get \n target value'+measure_mm+' in deg\n \n OFFSET',font = "Arial 10 bold italic").grid(row=5,column=0)
                for k in range (8):
                    k=k+1
                    if arm_num==str(k)+'2':    
                        lab=Label(extra,text='Guess the offset to get \n target value'+measure_mm+' in deg\n \n OFFSET',font = "Arial 10 bold italic").grid(row=5,column=0)
                    
                arm_selector = StringVar(extra)
                arm_selector.set(arm_num)
                arm_rad = Entry(master=extra,width=5)
                arm_rad.grid(row=6,column=0)
                ver_rad = Button(master = extra, command =lambda:confirm_rad(), height = 1, width =2, text = "☑️") 
                ver_rad.grid(row=7,column=0)        
                mainw.update_idletasks()
                frame_canvas.config(width=screen_width-10,
                                    height=screen_height-10)
                
                # Set the canvas scrolling region
                canvasf.config(scrollregion=canvasf.bbox("all"))               
                def confirm_rad():   
                    global arm_selector
                    global measure_selector 
                    global arm_num
                    global arm_rad
                    global arm_rad1    
                    Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)
                    Label(extra,text="Choose arm number",font = "Arial 10 bold italic").grid(row=1,column=0)
                    arm_selector = StringVar(extra)
                    arm_selector.set(arm_num)
                    filterop = OptionMenu(extra,arm_selector, *sorted({'11','12','21','22','31','32','41','42','51','52','61','62','71','72','81','82'}),)
                    filterop.grid(row=2, column=0)
                    measure_selector = StringVar(extra)
                    measure_selector.set('     Measurement holes     ')
                    measureop = OptionMenu(extra,measure_selector, *sorted({'third hole,second hole','third hole,first hole','second hole,first hole'}),)
                    measureop.grid(row=3, column=0)              
                    verty_inc = Button(master = extra, command =lambda:confirm_arm(), height = 2, width =10, text = "CONFIRM") 
                    verty_inc.grid(row=4,column=0) 
                    
                    arm_num=str(arm_selector.get()) 
                    try:
                        arm_rad=str(arm_rad.get())
                    except:
                        arm_rad=str(arm_rad1.get())
                    for k in range (8):
                        k=k+1
                        if arm_num==str(k)+'1':    
                            lab=Label(extra,text='Guess the offset to get \n target value'+measure_mm+' in deg\n \n OFFSET',font = "Arial 10 bold italic").grid(row=5,column=0)
                    for k in range (8):
                        k=k+1
                        if arm_num==str(k)+'2':    
                            lab=Label(extra,text='Guess the offset to get \n target value'+measure_mm+' in deg\n \n OFFSET',font = "Arial 10 bold italic").grid(row=5,column=0)
                        
                    arm_rad1 = Entry(master=extra,width=5)
                    arm_rad1.grid(row=6,column=0)
                    arm_rad1.insert(0,arm_rad)
                    ver_rad = Button(master = extra, command =lambda:confirm_rad2(), height = 1, width =8, text = "Double click") 
                    ver_rad.grid(row=7,column=0)
                    ver_rad = Button(master = extra, command =lambda:confirm_deg(), height = 1, width =8, text = "Save offset") 
                    ver_rad.grid(row=8,column=0)      
                    x_new_={}
                    y_new_={}
                    z_new_={}
                    calibration_angle=pd.read_excel(r'calibration.xlsx', engine='openpyxl')

                    for k in range (8):
                        k=k+1
                        x_new_[str(k)+'1']=np.zeros(numdata)
                        x_new_[str(k)+'2']=np.zeros(numdata)
                        y_new_[str(k)+'1']=np.zeros(numdata)
                        y_new_[str(k)+'2']=np.zeros(numdata)
                        z_new_[str(k)+'1']=np.zeros(numdata)
                        z_new_[str(k)+'2']=np.zeros(numdata)
                    try:
                        for k in range (8):
                            k=k+1
                            if arm_num==str(k)+'1':       
                                for i in range (numdata):
                                    theta=float(arm_rad)
                                    z_new_[str(k)+'1'][i]=((math.sin(math.radians(theta))*acc_sensor['y_'+str(k)+'1'][i])+(math.cos(math.radians(theta))*acc_sensor['z_'+str(k)+'1'][i]))
                                    y_new_[str(k)+'1'][i]=((math.cos(math.radians(theta))*acc_sensor['y_'+str(k)+'1'][i])-(math.sin(math.radians(theta))*acc_sensor['z_'+str(k)+'1'][i]))
                        for k in range (8):
                            k=k+1
                            if arm_num==str(k)+'2':       
                                for i in range (numdata):
                                    theta=float(arm_rad)
                                    z_new_[str(k)+'2'][i]=((math.sin(math.radians(theta))*acc_sensor['y_'+str(k)+'2'][i])+(math.cos(math.radians(theta))*acc_sensor['z_'+str(k)+'2'][i]))
                                    y_new_[str(k)+'2'][i]=((math.cos(math.radians(theta))*acc_sensor['y_'+str(k)+'2'][i])-(math.sin(math.radians(theta))*acc_sensor['z_'+str(k)+'2'][i]))

                        
                        a=np.zeros(numdata)
                        for i in range (numdata-1):
                            if int(arm_num[0])<5 and int(arm_num[1])==1:
                                try:
                                    degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i]=np.arccos(((acc_sensor['x_'+arm_num][i]*xarm_calibrated[arm_num[0]+'3'][i])+(y_new_[arm_num][i]*yarm_calibrated[arm_num[0]+'3'][i])+(z_new_[arm_num][i]*zarm_calibrated[arm_num[0]+'3'][i]))/(math.sqrt((acc_sensor['x_'+arm_num][i]**2)+(y_new_[arm_num][i]**2)+(z_new_[arm_num][i]**2))*(math.sqrt((xarm_calibrated[arm_num[0]+'3'][i]**2)+(yarm_calibrated[arm_num[0]+'3'][i]**2)+(zarm_calibrated[arm_num[0]+'3'][i]**2)))))*180/math.pi        
                                    a[i]=(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i])
                                except:
                                    a[i]=a[i-1]
                                    
                            if int(arm_num[0])<5 and int(arm_num[1])==2:  
                                try:
                                    degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i]=np.arccos(((-acc_sensor['x_'+arm_num][i]*xarm_calibrated[arm_num[0]+'3'][i])+(y_new_[arm_num][i]*yarm_calibrated[arm_num[0]+'3'][i])+(-z_new_[arm_num][i]*zarm_calibrated[arm_num[0]+'3'][i]))/(math.sqrt((acc_sensor['x_'+arm_num][i]**2)+(y_new_[arm_num][i]**2)+(z_new_[arm_num][i]**2))*(math.sqrt((xarm_calibrated[arm_num[0]+'3'][i]**2)+(yarm_calibrated[arm_num[0]+'3'][i]**2)+(zarm_calibrated[arm_num[0]+'3'][i]**2)))))*180/math.pi      
                                    a[i]=(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i])
                                except:
                                    a[i]=a[i-1]                                
                            if int(arm_num[0])>4 and int(arm_num[1])==1:  
                                try:
                                    degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i]=np.arccos(((acc_sensor['x_'+arm_num][i]*xarm_calibrated[arm_num[0]+'3'][i])+(-y_new_[arm_num][i]*yarm_calibrated[arm_num[0]+'3'][i])+(z_new_[arm_num][i]*zarm_calibrated[arm_num[0]+'3'][i]))/(math.sqrt((acc_sensor['x_'+arm_num][i]**2)+(y_new_[arm_num][i]**2)+(z_new_[arm_num][i]**2))*(math.sqrt((xarm_calibrated[arm_num[0]+'3'][i]**2)+(yarm_calibrated[arm_num[0]+'3'][i]**2)+(zarm_calibrated[arm_num[0]+'3'][i]**2)))))*180/math.pi      
                                    a[i]=(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i])
                                except:
                                    a[i]=a[i-1]                                
                            if int(arm_num[0])>4 and int(arm_num[1])==2:  
                                try:
                                    degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i]=180-np.arccos(((-acc_sensor['x_'+arm_num][i]*xarm_calibrated[arm_num[0]+'3'][i])+(-y_new_[arm_num][i]*yarm_calibrated[arm_num[0]+'3'][i])+(-z_new_[arm_num][i]*zarm_calibrated[arm_num[0]+'3'][i]))/(math.sqrt((acc_sensor['x_'+arm_num][i]**2)+(y_new_[arm_num][i]**2)+(z_new_[arm_num][i]**2))*(math.sqrt((xarm_calibrated[arm_num[0]+'3'][i]**2)+(yarm_calibrated[arm_num[0]+'3'][i]**2)+(zarm_calibrated[arm_num[0]+'3'][i]**2)))))*180/math.pi        
                                    a[i]=(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i]) 
                                except:
                                    a[i]=a[i-1]                                
                                
                        for i in range (numdata-1): 
                            if math.isnan(a[i])==True:
                                a[i]=a[i-1]
                        a[numdata-1]=a[numdata-2]
                        off=np.abs((np.abs(np.mean(a[0:num1]))-np.abs(np.mean(a[num1+1:numdata]))))
                        lab=Label(extra,text='According to your guess the angle difference is'+str(round(off,2)),font = "Arial 10 bold italic").grid(row=9,column=0)
                        mainw.update_idletasks()
                        frame_canvas.config(width=screen_width-10,
                                            height=screen_height-10)
                        
                        # Set the canvas scrolling region
                        canvasf.config(scrollregion=canvasf.bbox("all"))   
                        f=plt.figure()
                        for widget in plotplot.winfo_children():
                            widget.destroy()
                        canvas = FigureCanvasTkAgg(f, master =plotplot)
                        canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)
                
                        toolbarFrame = Frame(master=plotplot)
                        toolbarFrame.grid(row=1,column=0)
                        toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)
                
                        canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4) 
                        for widget in plotplot.winfo_children():
                            widget.destroy()        
                        canvas = Canvas(master =plotplot,bg='white',width='900',height='500')
                        canvas.grid(row=0,column=0, padx=4, pady=4)                             
                        f = plt.figure()
                        ax = f.add_subplot(111)
                        plt.plot(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'],label="Arm angle"+arm_num)
                        plt.legend(fontsize=8,loc='best')
                        plt.xlabel('Time (sec)')
                        plt.ylabel('Angle (deg)')

                        for widget in plotplot.winfo_children():
                            widget.destroy()   
                        canvas = FigureCanvasTkAgg(f, master =plotplot)
                        canvas.draw()
            
                        canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)
            
                        toolbarFrame = Frame(master=plotplot)
                        toolbarFrame.grid(row=1,column=0)
                        toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)
            
                        canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4)
                    except:
                        confirm_rad2()
                    def confirm_deg():
                        for k in range (8):
                            k=k+1
                            if arm_num==str(k)+'1':
                                sheet.cell(row=k*2,column=11).value=float(arm_rad)
                            if arm_num==str(k)+'2':
                                sheet.cell(row=(k*2)+1,column=11).value=float(arm_rad)
                            wb.save('calibration.xlsx')                         
                    def confirm_rad2(): 
                        global arm_selector
                        global arm_num
                        global arm_rad
                        ver_rad = Button(master = extra, command =lambda:confirm_rad(), height = 1, width =8, text = "Double click") 
                        ver_rad.grid(row=7,column=0)
                        mainw.update_idletasks()
                        frame_canvas.config(width=screen_width-10,
                                            height=screen_height-10)
                        
                        # Set the canvas scrolling region
                        canvasf.config(scrollregion=canvasf.bbox("all"))
                        
            for i in range(numdata):
                time[i]=(0.5*i)
    
#################################ARM ANGLE STEP2##############################
    if option=='Arm radius step2':
        try:
            for widget in extra.winfo_children():
                    widget.destroy() 
        except:
            pass
        try:
            for widget in extra1.winfo_children():
                    widget.destroy() 
        except:
            pass  
                
        check_butt = Button(master = mainw, command =lambda:run_arm(), height = 1, width =10, text = "Load data") 
        check_butt.grid(row=3,column=0)
        def run_arm():
            global filename
            global numdata
            global num1
            global extra
            try:
                num1=numdata
            except:
                pass
            ##Load Data
            import load
            data,numdata,empty,NaN=load.l()
            ##############################################################################
            #Import all Variables
            
            import newvar as allvar
            alldata,acc_sensor,ii,\
            sensor_name,encoder_finder,encoder_diff,pressure,voltage,depth,aa,\
            thetadict,angleall,degreedict,degreeall,ljoint,larm,lcont,radius,\
            radiusdict,radiusall,angle,degree,inclinationdirection,incline,\
            magnetic_declination,inclination_x,inclination_y,anglecaliball,degreecaliball,\
            radiuscaliball,thetadict_calib,degree_calibdict,radius_calibdict,\
            angle_calibrated,degree_calibrated,radius_calibrated\
            =allvar.allvariable(numdata,data,empty,NaN)
            ###############################################################################
            #Accelerometer sensors
            
            import newacc as newacc
            acc_sensor=newacc.acc(numdata,data,alldata,ii,sensor_name,acc_sensor,\
              NaN,encoder_finder,encoder_diff)
            ###############################################################################
            #MAGNETOMETER SESNOR
            
            import newmag as mag_sensor
            magnetic_x,magnetic_y,magnetic_z,magneticfield_x_uncalibrated,\
            magneticfield_y_uncalibrated,magneticfield_z_uncalibrated,absolute_uncalibrated_magneticfield\
            =mag_sensor.magnetometer_sensor(numdata,data,NaN,encoder_finder,encoder_diff)
            ##############################################################################
            ##GYROSCOPE SENSOR
            
            import newgyro as gyro_sensor
            gyro_x,gyro_y,gyro_z,gyro_xuncalibrated,gyro_yuncalibrated,gyro_zuncalibrated,\
            gyro_xcalibrated,gyro_ycalibrated,gyro_zcalibrated,gyro_rotation_uncalibrated,gyro_rotation_calibrated,\
            gyro_rotation_xuncalibrated,gyro_rotation_xcalibrated,gyro_rotation_yuncalibrated,\
            gyro_rotation_ycalibrated\
            =gyro_sensor.gyroscope(numdata,data,NaN,encoder_finder,encoder_diff)
            ###############################################################################
            #Calculate the uncalibrated angle and radius of the arms
            
            import newangleradius   
            degreeall,radiusall\
            =newangleradius.ang_rad(numdata, angle, degree, aa, thetadict,\
            angleall, degreedict, degreeall, ljoint, larm, lcont, radius,\
            radiusdict, radiusall,acc_sensor)        
            ###############################################################################   
            #Uncalibrated rotation of the logger   
            
            import newcalc 
            sum_mag_declination,inclinationdirection\
            =newcalc.inclination(numdata,inclinationdirection,incline,\
            magnetic_declination,acc_sensor,magnetic_x,magnetic_y,magnetic_z)    
            ############################################################################### 
            #Inclination 
            
            import newinc
            inclination_x,inclination_y,xarm_calibrated,yarm_calibrated,zarm_calibrated,\
            xcalibrated,ycalibrated,zcalibrated,xx_calibrated,yy_calibrated,zz_calibrated,\
            inclinationdirection,incline,x3_avg,y3_avg,z3_avg,inclination_xaxis_avg,inclination_yaxis_avg\
            =newinc.calib(numdata,acc_sensor,sum_mag_declination,\
            inclination_x,inclination_y)
            ###############################################################################
            ##ARM CALIBRATION
            
            import newarmcalib
            acc_sensor,radius_calibdict,degree_calibdict\
            =newarmcalib.arm(acc_sensor,xarm_calibrated,yarm_calibrated,zarm_calibrated,\
            numdata,anglecaliball,degreecaliball,radiuscaliball,thetadict_calib,degree_calibdict,radius_calibdict,\
            angle_calibrated,degree_calibrated,radius_calibrated,aa,lcont,larm,ljoint) 
                
            global extra
            
            try:
                for widget in extra.winfo_children():
                        widget.destroy()
            except:
                pass
            clean()
            extra = Frame(mainw)
            extra.grid(row=5,column=1, sticky='W')                  
            disc = LabelFrame(mainw, text="Arm radius:", font="Arial 12 bold italic")
            disc.grid(row=4,column=0)
            scrollbar = Scrollbar(disc)
            discrip = tk.Text(disc, height=5, width=70, yscrollcommand=scrollbar.set, font="Arial 14")
            scrollbar.config(command=discrip.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            discrip.pack(side="left")
            discrip.insert(END,'degree arm_11= '+str(round(np.mean(degree_calibdict['degreecalib_11_13'][0:numdata-1]),2))+'  deg'+'\ndegree arm_21= '+str(round(np.mean(degree_calibdict['degreecalib_21_23'][0:numdata-1]),2))+'  deg'+'\ndegree arm_31= '+str(round(np.mean(degree_calibdict['degreecalib_31_33'][0:numdata-1]),2))+'  deg'+'\ndegree arm_41= '+str(round(np.mean(degree_calibdict['degreecalib_41_43'][0:numdata-1]),2))+'  deg'+'\ndegree arm_51= '+str(round(np.mean(degree_calibdict['degreecalib_51_53'][0:numdata-1]),2))+'  deg'+'\ndegree arm_61= '+str(round(np.mean(degree_calibdict['degreecalib_61_63'][0:numdata-1]),2))+'  deg'+'\ndegree arm_71= '+str(round(np.mean(degree_calibdict['degreecalib_71_73'][0:numdata-1]),2))+'  deg'+'\ndegree arm_81= '+str(round(np.mean(degree_calibdict['degreecalib_81_83'][0:numdata-1]),2))+'  deg'+'\ndegree arm_12= '+str(round(np.mean(degree_calibdict['degreecalib_12_13'][0:numdata-1]),2))+'  deg'+'\ndegree arm_22= '+str(round(np.mean(degree_calibdict['degreecalib_22_23'][0:numdata-1]),2))+'  deg'+'\ndegree arm_32= '+str(round(np.mean(degree_calibdict['degreecalib_32_33'][0:numdata-1]),2))+'  deg'+'\ndegree arm_42= '+str(round(np.mean(degree_calibdict['degreecalib_42_43'][0:numdata-1]),2))+'  deg'+'\ndegree arm_52= '+str(round(np.mean(degree_calibdict['degreecalib_52_53'][0:numdata-1]),2))+'  deg'+'\ndegree arm_62= '+str(round(np.mean(degree_calibdict['degreecalib_62_63'][0:numdata-1]),2))+'  deg'+'\ndegree arm_72= '+str(round(np.mean(degree_calibdict['degreecalib_72_73'][0:numdata-1]),2))+'  deg'+'\ndegree arm_82= '+str(round(np.mean(degree_calibdict['degreecalib_82_83'][0:numdata-1]),2))+'  deg')      
            plotplot = LabelFrame(mainw)
            plotplot.grid(row=5,column=0)
            time=np.zeros(numdata)
            Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)
            Label(extra,text="Choose arm number",font = "Arial 10 bold italic").grid(row=1,column=0)
            global arm_selector 
            global measure_selector 
            arm_selector = StringVar(extra)
            arm_selector.set(' Arm number ')
            filterop = OptionMenu(extra,arm_selector, *sorted({'11','12','21','22','31','32','41','42','51','52','61','62','71','72','81','82'}),)
            filterop.grid(row=2, column=0)
            measure_selector = StringVar(extra)
            measure_selector.set('     Measurement holes     ')
            measureop = OptionMenu(extra,measure_selector, *sorted({'third hole','first hole','second hole'}),)
            measureop.grid(row=3, column=0)
            
            verty_inc = Button(master = extra, command =lambda:confirm_arm(), height = 2, width =10, text = "CONFIRM") 
            verty_inc.grid(row=4,column=0)
            acc_sensor=newacc.acc(numdata,data,alldata,ii,sensor_name,acc_sensor,\
              NaN,encoder_finder,encoder_diff)
            mainw.update_idletasks()
            frame_canvas.config(width=screen_width-10,
                                height=screen_height-10)
            
            # Set the canvas scrolling region
            canvasf.config(scrollregion=canvasf.bbox("all"))            
            def confirm_arm():
                global arm_selector
                global measure_selector 
                global arm_num
                global measure_hole
                global measure_mm
                global arm_rad
                global extra
              
                for widget in extra.winfo_children():
                    widget.destroy()
                extra = Frame(mainw)
                extra.grid(row=5,column=1, sticky='W')  
                Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)                    
                Label(extra,text="Choose arm number",font = "Arial 10 bold italic").grid(row=1,column=0)
                global arm_selector 

                filterop = OptionMenu(extra,arm_selector, *sorted({'11','12','21','22','31','32','41','42','51','52','61','62','71','72','81','82'}),)
                filterop.grid(row=2, column=0)
                measureop = OptionMenu(extra,measure_selector, *sorted({'third hole','first hole','second hole'}),)
                measureop.grid(row=3, column=0)                
                verty_inc = Button(master = extra, command =lambda:confirm_arm(), height = 2, width =10, text = "CONFIRM") 
                verty_inc.grid(row=4,column=0)

                arm_num=str(arm_selector.get())
                measure_hole=str(measure_selector.get())
                if measure_hole=='third hole':
                    measure_mm=str(21.4)
                if measure_hole=='second hole':
                    measure_mm=str(14.45)
                if measure_hole=='first hole':
                    measure_mm=str(7.72)
                for k in range (8):
                    k=k+1
                    if arm_num==str(k)+'1':    
                        lab=Label(extra,text='Guess the offset to get \n target value'+measure_mm+' in deg\n \n OFFSET',font = "Arial 10 bold italic").grid(row=5,column=0)
                for k in range (8):
                    k=k+1
                    if arm_num==str(k)+'2':    
                        lab=Label(extra,text='Guess the offset to get \n target value'+measure_mm+' in deg\n \n OFFSET',font = "Arial 10 bold italic").grid(row=5,column=0)
                
                arm_selector = StringVar(extra)
                arm_selector.set(arm_num)
                arm_rad = Entry(master=extra,width=5)
                arm_rad.grid(row=6,column=0)
                ver_rad = Button(master = extra, command =lambda:confirm_rad(), height = 1, width =2, text = "☑️") 
                ver_rad.grid(row=7,column=0)
                mainw.update_idletasks()
                frame_canvas.config(width=screen_width-10,
                                    height=screen_height-10)
                
                # Set the canvas scrolling region
                canvasf.config(scrollregion=canvasf.bbox("all"))                                
                def confirm_rad():   
                    global arm_selector
                    global measure_selector 
                    global arm_num
                    global arm_rad
                    global arm_rad1   
                    Label(extra,text="Validation",font = "Arial 10 italic").grid(row=0,column=0)
                    Label(extra,text="Choose arm number",font = "Arial 10 bold italic").grid(row=1,column=0)
                    arm_selector = StringVar(extra)
                    arm_selector.set(arm_num)
                    filterop = OptionMenu(extra,arm_selector, *sorted({'11','12','21','22','31','32','41','42','51','52','61','62','71','72','81','82'}),)
                    filterop.grid(row=2, column=0)
                    measure_selector = StringVar(extra)
                    measure_selector.set('     Measurement holes     ')
                    measureop = OptionMenu(extra,measure_selector, *sorted({'third hole','first hole','second hole'}),)
                    measureop.grid(row=3, column=0)              
                    verty_inc = Button(master = extra, command =lambda:confirm_arm(), height = 2, width =10, text = "CONFIRM") 
                    verty_inc.grid(row=4,column=0) 
                    
                    arm_num=str(arm_selector.get()) 
                    try:
                        arm_rad=str(arm_rad.get())
                    except:
                        arm_rad=str(arm_rad1.get())
                    for k in range (8):
                        k=k+1
                        if arm_num==str(k)+'1':    
                            lab=Label(extra,text='Guess the offset to get \n target value'+measure_mm+' in deg\n \n OFFSET',font = "Arial 10 bold italic").grid(row=5,column=0)
                    for k in range (8):
                        k=k+1
                    if arm_num==str(k)+'2':    
                        lab=Label(extra,text='Guess the offset to get \n target value'+measure_mm+' in deg\n \n OFFSET',font = "Arial 10 bold italic").grid(row=5,column=0)
                    arm_rad1 = Entry(master=extra,width=5)
                    arm_rad1.grid(row=6,column=0)
                    arm_rad1.insert(0,arm_rad)
                    ver_rad = Button(master = extra, command =lambda:confirm_rad2(), height = 1, width =8, text = "Double click") 
                    ver_rad.grid(row=7,column=0)
                    ver_rad = Button(master = extra, command =lambda:confirm_deg(), height = 1, width =8, text = "Save offset") 
                    ver_rad.grid(row=8,column=0)                    
                    
                    x_new_={}
                    y_new_={}
                    z_new_={}
                    for k in range (8):
                        k=k+1
                        x_new_[str(k)+'1']=np.zeros(numdata)
                        x_new_[str(k)+'2']=np.zeros(numdata)
                        y_new_[str(k)+'1']=np.zeros(numdata)
                        y_new_[str(k)+'2']=np.zeros(numdata)
                        z_new_[str(k)+'1']=np.zeros(numdata)
                        z_new_[str(k)+'2']=np.zeros(numdata)
    
                    calibration_angle=pd.read_excel(r'calibration.xlsx', engine='openpyxl')
                    
                    try:
                        for k in range (8):
                            k=k+1
                            if arm_num==str(k)+'1':       
                                for i in range (numdata):
                                    theta=float(arm_rad)
                                    y_new_[str(k)+'1'][i]=((math.sin(math.radians(theta))*acc_sensor['x_'+str(k)+'1'][i])+(math.cos(math.radians(theta))*acc_sensor['y_'+str(k)+'1'][i]))
                                    x_new_[str(k)+'1'][i]=((math.cos(math.radians(theta))*acc_sensor['x_'+str(k)+'1'][i])-(math.sin(math.radians(theta))*acc_sensor['y_'+str(k)+'1'][i]))
                                    theta=np.array(calibration_angle['offset_along_x'][k-1])
                                    z_new_[str(k)+'1'][i]=((math.sin(math.radians(theta))*y_new_[str(k)+'1'][i])+(math.cos(math.radians(theta))*acc_sensor['z_'+str(k)+'1'][i]))
                                    y_new_[str(k)+'1'][i]=((math.cos(math.radians(theta))*y_new_[str(k)+'1'][i])-(math.sin(math.radians(theta))*acc_sensor['z_'+str(k)+'1'][i]))

                        for k in range (8):
                            k=k+1
                            if arm_num==str(k)+'2':       
                                for i in range (numdata):
                                    theta=float(arm_rad)
                                    y_new_[str(k)+'2'][i]=((math.sin(math.radians(theta))*acc_sensor['x_'+str(k)+'2'][i])+(math.cos(math.radians(theta))*acc_sensor['y_'+str(k)+'2'][i]))
                                    x_new_[str(k)+'2'][i]=((math.cos(math.radians(theta))*acc_sensor['x_'+str(k)+'2'][i])-(math.sin(math.radians(theta))*acc_sensor['y_'+str(k)+'2'][i]))
                                    theta=np.array(calibration_angle['offset_along_x'][(k*2)+1])
                                    z_new_[str(k)+'2'][i]=((math.sin(math.radians(theta))*y_new_[str(k)+'2'][i])+(math.cos(math.radians(theta))*acc_sensor['z_'+str(k)+'2'][i]))
                                    y_new_[str(k)+'2'][i]=((math.cos(math.radians(theta))*y_new_[str(k)+'2'][i])-(math.sin(math.radians(theta))*acc_sensor['z_'+str(k)+'2'][i]))
                    
                        
                        a=np.zeros(numdata)
                        r=np.zeros(numdata)
                        for i in range (numdata):
                            if int(arm_num[0])<5 and int(arm_num[1])==1:
                                degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i]=np.arccos(((x_new_[arm_num][i]*xarm_calibrated[arm_num[0]+'3'][i])+(y_new_[arm_num][i]*yarm_calibrated[arm_num[0]+'3'][i])+(z_new_[arm_num][i]*zarm_calibrated[arm_num[0]+'3'][i]))/(math.sqrt((x_new_[arm_num][i]**2)+(y_new_[arm_num][i]**2)+(z_new_[arm_num][i]**2))*(math.sqrt((xarm_calibrated[arm_num[0]+'3'][i]**2)+(yarm_calibrated[arm_num[0]+'3'][i]**2)+(zarm_calibrated[arm_num[0]+'3'][i]**2)))))*180/math.pi  
                                a[i]=(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i])
                            if int(arm_num[0])<5 and int(arm_num[1])==2:  
                                degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i]=np.arccos(((-x_new_[arm_num][i]*xarm_calibrated[arm_num[0]+'3'][i])+(y_new_[arm_num][i]*yarm_calibrated[arm_num[0]+'3'][i])+(-z_new_[arm_num][i]*zarm_calibrated[arm_num[0]+'3'][i]))/(math.sqrt((x_new_[arm_num][i]**2)+(y_new_[arm_num][i]**2)+(z_new_[arm_num][i]**2))*(math.sqrt((xarm_calibrated[arm_num[0]+'3'][i]**2)+(yarm_calibrated[arm_num[0]+'3'][i]**2)+(zarm_calibrated[arm_num[0]+'3'][i]**2)))))*180/math.pi                   
                                a[i]=(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i])
                            if int(arm_num[0])>4 and int(arm_num[1])==1:  
                                degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i]=np.arccos(((-x_new_[arm_num][i]*xarm_calibrated[arm_num[0]+'3'][i])+(y_new_[arm_num][i]*yarm_calibrated[arm_num[0]+'3'][i])+(-z_new_[arm_num][i]*zarm_calibrated[arm_num[0]+'3'][i]))/(math.sqrt((x_new_[arm_num][i]**2)+(y_new_[arm_num][i]**2)+(z_new_[arm_num][i]**2))*(math.sqrt((xarm_calibrated[arm_num[0]+'3'][i]**2)+(yarm_calibrated[arm_num[0]+'3'][i]**2)+(zarm_calibrated[arm_num[0]+'3'][i]**2)))))*180/math.pi  
                                a[i]=(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i])
                            if int(arm_num[0])>4 and int(arm_num[1])==2:  
                                degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i]=np.arccos(((x_new_[arm_num][i]*xarm_calibrated[arm_num[0]+'3'][i])+(-y_new_[arm_num][i]*yarm_calibrated[arm_num[0]+'3'][i])+(z_new_[arm_num][i]*zarm_calibrated[arm_num[0]+'3'][i]))/(math.sqrt((x_new_[arm_num][i]**2)+(y_new_[arm_num][i]**2)+(z_new_[arm_num][i]**2))*(math.sqrt((xarm_calibrated[arm_num[0]+'3'][i]**2)+(yarm_calibrated[arm_num[0]+'3'][i]**2)+(zarm_calibrated[arm_num[0]+'3'][i]**2)))))*180/math.pi  
                                a[i]=(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'][i])    
                        for i in range (numdata):
                            if math.isnan(a[i])==True:
                                a[i]=a[i-1]                                
                        for i in range (numdata): 
                                radius_calibdict['radiuscalib_'+arm_num+'_'+arm_num[0]+'3'][i]=ljoint+lcont+(larm*(math.sin(math.radians(a[i]))))     
                                r[i]=radius_calibdict['radiuscalib_'+arm_num+'_'+arm_num[0]+'3'][i]

                          

                        off=(np.mean(a))
                        lab=Label(extra,text='According to your guess the arm angle is'+str(round(off,2)),font = "Arial 10 bold italic").grid(row=9,column=0)
                        mainw.update_idletasks()
                        frame_canvas.config(width=screen_width-10,
                                            height=screen_height-10)
                        
                        # Set the canvas scrolling region
                        canvasf.config(scrollregion=canvasf.bbox("all"))   
                        f=plt.figure()
                        for widget in plotplot.winfo_children():
                            widget.destroy()
                        canvas = FigureCanvasTkAgg(f, master =plotplot)
                        canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)
                
                        toolbarFrame = Frame(master=plotplot)
                        toolbarFrame.grid(row=1,column=0)
                        toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)
                
                        canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4) 
                        for widget in plotplot.winfo_children():
                            widget.destroy()        
                        canvas = Canvas(master =plotplot,bg='white',width='900',height='500')
                        canvas.grid(row=0,column=0, padx=4, pady=4)                             
                        f = plt.figure()
                        ax = f.add_subplot(111)
                        plt.plot(degree_calibdict['degreecalib_'+arm_num+'_'+arm_num[0]+'3'],label="Arm angle"+arm_num)
                        plt.legend(fontsize=8,loc='best')
                        plt.xlabel('Time (sec)')
                        plt.ylabel('Angle (deg)')
            
                        for widget in plotplot.winfo_children():
                            widget.destroy()   
                        canvas = FigureCanvasTkAgg(f, master =plotplot)
                        canvas.draw()
            
                        canvas.get_tk_widget().grid(row=0,column=0, padx=4, pady=4)
            
                        toolbarFrame = Frame(master=plotplot)
                        toolbarFrame.grid(row=1,column=0)
                        toolbar = NavigationToolbar2Tk(canvas, toolbarFrame)
            
                        canvas._tkcanvas.grid(row=2,column=0, padx=4, pady=4)    

                    except:
                        confirm_rad2()
                    def confirm_deg():
                        for k in range (8):
                            k=k+1
                            if arm_num==str(k)+'1':
                                sheet.cell(row=k*2,column=12).value=float(arm_rad)
                            if arm_num==str(k)+'2':
                                sheet.cell(row=(k*2)+1,column=12).value=float(arm_rad)
                            wb.save('calibration.xlsx')                         
                        
                    def confirm_rad2(): 
                        global arm_selector
                        global arm_num
                        global arm_rad
                        ver_rad = Button(master = extra, command =lambda:confirm_rad(), height = 1, width =8, text = "Double click") 
                        ver_rad.grid(row=7,column=0)
                        mainw.update_idletasks()
                        frame_canvas.config(width=screen_width-10,
                                            height=screen_height-10)
                        
                        # Set the canvas scrolling region
                        canvasf.config(scrollregion=canvasf.bbox("all"))
                        
                # arm_rad=float(y_inc.get())                
            for i in range(numdata):
                time[i]=(0.5*i)


mainloop()